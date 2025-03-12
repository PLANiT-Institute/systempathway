import pandas as pd
import os
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, Series

def merge_excel_files(input_files, output_file, add_colors=True):
    """
    Merge multiple Excel files into a single Excel file with scenario identifiers.
    
    Parameters:
    input_files (list): List of input Excel file paths
    output_file (str): Path for the output Excel file
    add_colors (bool): Whether to add color coding to the merged file
    """
    # Check if any input files exist
    valid_files = [f for f in input_files if os.path.exists(f)]
    if not valid_files:
        print("Error: None of the specified input files exist.")
        return False
    
    # Create a new Excel writer
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    
    # Track which sheets we've seen
    processed_sheets = set()
    
    # Track scenarios found
    scenarios_found = []
    
    # Process each input file
    for file_idx, file_path in enumerate(input_files):
        if not os.path.exists(file_path):
            print(f"Warning: File {file_path} does not exist. Skipping...")
            continue
            
        # Extract scenario name from filename (e.g., 'g1' from 'results_g1.xlsx')
        scenario = os.path.basename(file_path).split('.')[0].split('_')[-1]
        scenarios_found.append(scenario)
        print(f"Processing {file_path} (Scenario: {scenario})...")
        
        try:
            # Read each sheet from the current Excel file
            excel_file = pd.ExcelFile(file_path)
            
            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    # Add a scenario column to identify the source file
                    df['Scenario'] = scenario
                    
                    # Move Scenario column to the beginning
                    cols = df.columns.tolist()
                    cols = cols[-1:] + cols[:-1]
                    df = df[cols]
                    
                    # If we've already processed this sheet type, append to existing sheet
                    merged_sheet_name = sheet_name
                    if sheet_name in processed_sheets:
                        # Read existing data
                        existing_df = pd.read_excel(output_file, sheet_name=merged_sheet_name)
                        # Combine with new data
                        combined_df = pd.concat([existing_df, df], ignore_index=True)
                        # Write back to the same sheet
                        combined_df.to_excel(writer, sheet_name=merged_sheet_name, index=False)
                    else:
                        # First time seeing this sheet, just write it
                        df.to_excel(writer, sheet_name=merged_sheet_name, index=False)
                        processed_sheets.add(sheet_name)
                except Exception as e:
                    print(f"Error processing sheet {sheet_name} in {file_path}: {str(e)}")
        except Exception as e:
            print(f"Error processing file {file_path}: {str(e)}")
    
    # Save the output file
    writer.close()
    print(f"Merged data saved to {output_file}")
    
    # Add color-coding and formatting
    if add_colors:
        add_color_coding(output_file, scenarios_found)
    
    # Create summary sheet with charts
    create_summary_charts(output_file, scenarios_found)
    
    return True

def add_color_coding(excel_file, scenarios=None):
    """
    Add color-coding to an Excel file based on scenario values.
    
    Parameters:
    excel_file (str): Path to the Excel file to modify
    scenarios (list): List of scenarios found in the data
    """
    # Define colors for different scenarios (can handle more scenarios)
    scenario_colors = {
        'g1': 'FFCCFFCC',  # Light green
        'g2': 'FFFFCCCC',  # Light red/pink
        'g3': 'FFCCCCFF',  # Light blue
        'g4': 'FFFFF2CC',  # Light yellow
        'g5': 'FFE6E6E6',  # Light gray
        'g6': 'FFD9D2E9',  # Light purple
        'g7': 'FFDDEEFF',  # Light cyan
        'g8': 'FFFCE5CD',  # Light orange
    }
    
    # Define header style
    header_fill = PatternFill(start_color="FF366092", end_color="FF366092", fill_type='solid')
    header_font = Font(color="FFFFFFFF", bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Load the workbook
    try:
        wb = load_workbook(excel_file)
        
        # Process each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) if max_length < 50 else 50
                ws.column_dimensions[column].width = adjusted_width
            
            # Find the scenario column index
            scenario_col = None
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value == 'Scenario':
                    scenario_col = col_idx
                    break
            
            if scenario_col is None:
                continue  # Skip if no Scenario column
                
            # Apply colors based on scenario value
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                cell = row[scenario_col-1]  # Convert to 0-based index
                if cell.value in scenario_colors:
                    fill = PatternFill(start_color=scenario_colors[cell.value], 
                                      end_color=scenario_colors[cell.value], 
                                      fill_type='solid')
                    for c in row:
                        c.fill = fill
                        c.border = thin_border
        
        # Save the workbook
        wb.save(excel_file)
        print(f"Added color coding and formatting to {excel_file}")
    except Exception as e:
        print(f"Error adding color coding: {str(e)}")

def create_summary_charts(excel_file, scenarios):
    """
    Create a summary sheet with charts comparing key metrics across scenarios.
    
    Parameters:
    excel_file (str): Path to the Excel file
    scenarios (list): List of scenarios to include in the charts
    """
    try:
        # Load the workbook
        wb = load_workbook(excel_file)
        
        # Check if Global_Summary sheet exists
        if 'Global_Summary' not in wb.sheetnames:
            print("No Global_Summary sheet found. Skipping chart creation.")
            return
        
        # Create a new summary sheet
        if 'Scenario_Comparison' in wb.sheetnames:
            ws_summary = wb['Scenario_Comparison']
            # Clear existing content
            for row in ws_summary.iter_rows():
                for cell in row:
                    cell.value = None
        else:
            ws_summary = wb.create_sheet('Scenario_Comparison', 0)
        
        # Set title
        ws_summary['A1'] = 'Scenario Comparison Charts'
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary.merge_cells('A1:H1')
        
        # Read Global_Summary data
        ws_global = wb['Global_Summary']
        
        # Find key metrics columns
        header_row = next(ws_global.rows)
        col_indices = {}
        for idx, cell in enumerate(header_row):
            col_indices[cell.value] = idx
        
        # Key metrics to chart
        metrics = [
            'Total_Emissions', 
            'Total_Production', 
            'Global_Emission_Intensity',
            'Total_CAPEX',
            'Total_OPEX',
            'Total_Cost'
        ]
        
        # Create data for each metric
        row_offset = 3
        for metric_idx, metric in enumerate(metrics):
            if metric not in col_indices:
                continue
                
            # Add metric title
            title_row = row_offset + metric_idx * 20
            ws_summary.cell(row=title_row, column=1, value=f"{metric} by Scenario and Year")
            ws_summary.cell(row=title_row, column=1).font = Font(size=14, bold=True)
            ws_summary.merge_cells(f'A{title_row}:H{title_row}')
            
            # Add headers
            ws_summary.cell(row=title_row+1, column=1, value="Year")
            for scen_idx, scenario in enumerate(scenarios):
                ws_summary.cell(row=title_row+1, column=2+scen_idx, value=scenario)
            
            # Extract data
            years = set()
            data = {scenario: {} for scenario in scenarios}
            
            for row in ws_global.iter_rows(min_row=2):
                year = row[col_indices['Year']].value
                scenario = row[col_indices['Scenario']].value
                if scenario in scenarios and metric in col_indices:
                    value = row[col_indices[metric]].value
                    if value is not None:
                        years.add(year)
                        data[scenario][year] = value
            
            # Sort years
            years = sorted(years)
            
            # Write data
            for year_idx, year in enumerate(years):
                ws_summary.cell(row=title_row+2+year_idx, column=1, value=year)
                for scen_idx, scenario in enumerate(scenarios):
                    if year in data[scenario]:
                        ws_summary.cell(row=title_row+2+year_idx, column=2+scen_idx, value=data[scenario][year])
            
            # Create chart
            chart = LineChart()
            chart.title = f"{metric} by Year and Scenario"
            chart.style = 2
            chart.x_axis.title = "Year"
            chart.y_axis.title = metric.replace('_', ' ')
            
            data_ref = Reference(ws_summary, min_row=title_row+1, max_row=title_row+1+len(years),
                                min_col=1, max_col=1+len(scenarios))
            cats_ref = Reference(ws_summary, min_row=title_row+2, max_row=title_row+1+len(years),
                                min_col=1, max_col=1)
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            
            # Place chart
            ws_summary.add_chart(chart, f"J{title_row}")
        
        # Save workbook
        wb.save(excel_file)
        print(f"Created summary charts in {excel_file}")
    except Exception as e:
        print(f"Error creating summary charts: {str(e)}")

def compare_scenarios(input_files, output_file, metrics_to_compare=None):
    """
    Create a focused comparison of specific metrics across scenarios.
    
    Parameters:
    input_files (list): List of input Excel file paths
    output_file (str): Path for the output Excel file
    metrics_to_compare (list): List of metrics to compare (if None, uses default key metrics)
    """
    if metrics_to_compare is None:
        metrics_to_compare = [
            'Total_Emissions', 
            'Total_Production', 
            'Global_Emission_Intensity',
            'Total_CAPEX',
            'Total_OPEX',
            'Total_Cost'
        ]
    
    # First merge the files
    if not merge_excel_files(input_files, output_file):
        return
    
    # Now create a focused comparison sheet
    try:
        # Load the workbook
        wb = load_workbook(output_file)
        
        # Check if Global_Summary sheet exists
        if 'Global_Summary' not in wb.sheetnames:
            print("No Global_Summary sheet found. Cannot create comparison.")
            return
        
        # Create a new comparison sheet
        if 'Metric_Comparison' in wb.sheetnames:
            ws_compare = wb['Metric_Comparison']
            # Clear existing content
            for row in ws_compare.iter_rows():
                for cell in row:
                    cell.value = None
        else:
            ws_compare = wb.create_sheet('Metric_Comparison')
        
        # Set title
        ws_compare['A1'] = 'Scenario Metric Comparison'
        ws_compare['A1'].font = Font(size=16, bold=True)
        ws_compare.merge_cells('A1:F1')
        
        # Read Global_Summary data
        df = pd.read_excel(output_file, sheet_name='Global_Summary')
        
        # Get unique scenarios and years
        scenarios = df['Scenario'].unique().tolist()
        years = df['Year'].unique().tolist()
        years.sort()
        
        # For each metric, create a comparison table
        row_offset = 3
        for metric_idx, metric in enumerate(metrics_to_compare):
            if metric not in df.columns:
                continue
                
            # Add metric title
            title_row = row_offset + metric_idx * (len(years) + 5)
            ws_compare.cell(row=title_row, column=1, value=f"{metric.replace('_', ' ')} Comparison")
            ws_compare.cell(row=title_row, column=1).font = Font(size=14, bold=True)
            ws_compare.merge_cells(f'A{title_row}:F{title_row}')
            
            # Add headers
            ws_compare.cell(row=title_row+1, column=1, value="Year")
            for scen_idx, scenario in enumerate(scenarios):
                ws_compare.cell(row=title_row+1, column=2+scen_idx, value=scenario)
            
            # Add data rows
            for year_idx, year in enumerate(years):
                ws_compare.cell(row=title_row+2+year_idx, column=1, value=year)
                
                for scen_idx, scenario in enumerate(scenarios):
                    # Get value for this scenario and year
                    value = df[(df['Scenario'] == scenario) & (df['Year'] == year)][metric].values
                    if len(value) > 0:
                        ws_compare.cell(row=title_row+2+year_idx, column=2+scen_idx, value=value[0])
            
            # Add percent difference row (compared to first scenario)
            if len(scenarios) > 1:
                diff_row = title_row+2+len(years)
                ws_compare.cell(row=diff_row, column=1, value="% Diff from " + scenarios[0])
                
                for scen_idx, scenario in enumerate(scenarios[1:], 1):
                    # Calculate average percent difference across years
                    diffs = []
                    for year_idx, year in enumerate(years):
                        base_val = df[(df['Scenario'] == scenarios[0]) & (df['Year'] == year)][metric].values
                        comp_val = df[(df['Scenario'] == scenario) & (df['Year'] == year)][metric].values
                        
                        if len(base_val) > 0 and len(comp_val) > 0 and base_val[0] != 0:
                            diff_pct = (comp_val[0] - base_val[0]) / base_val[0] * 100
                            diffs.append(diff_pct)
                    
                    if diffs:
                        avg_diff = sum(diffs) / len(diffs)
                        ws_compare.cell(row=diff_row, column=2+scen_idx, value=avg_diff)
                        # Format as percentage
                        ws_compare.cell(row=diff_row, column=2+scen_idx).number_format = '0.00%'
            
            # Create chart
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = f"{metric.replace('_', ' ')} by Scenario"
            chart.y_axis.title = metric.replace('_', ' ')
            chart.x_axis.title = "Year"
            
            data_len = len(years)
            data = Reference(ws_compare, min_row=title_row+1, max_row=title_row+1+data_len,
                           min_col=2, max_col=1+len(scenarios))
            cats = Reference(ws_compare, min_row=title_row+2, max_row=title_row+1+data_len,
                           min_col=1, max_col=1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # Place chart
            ws_compare.add_chart(chart, f"H{title_row}")
        
        # Save workbook
        wb.save(output_file)
        print(f"Created metric comparison in {output_file}")
    except Exception as e:
        print(f"Error creating metric comparison: {str(e)}")

if __name__ == "__main__":
    # Input files to merge
    input_files = [
        'results_g1.xlsx',
        'results_g2.xlsx',
        'results_g3.xlsx'
    ]
    
    # Output file
    output_file = 'merged_results.xlsx'
    
    # Merge the files
    merge_excel_files(input_files, output_file)
    
    # Create a focused comparison of key metrics
    key_metrics = [
        'Total_Emissions', 
        'Total_Production', 
        'Global_Emission_Intensity',
        'Total_CAPEX',
        'Total_OPEX',
        'Total_Cost'
    ]
    
    compare_scenarios(input_files, 'scenario_comparison.xlsx', key_metrics)
