import pandas as pd
import os
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference, Series

def merge_excel_files(input_files, output_file, add_colors=True, calculate_fuel_costs=True):
    """
    Merge multiple Excel files into a single Excel file with scenario identifiers.
    
    Parameters:
    input_files (list): List of input Excel file paths
    output_file (str): Path for the output Excel file
    add_colors (bool): Whether to add color coding to the merged file
    calculate_fuel_costs (bool): Whether to calculate and add fuel costs
    """
    # Check if any input files exist
    valid_files = [f for f in input_files if os.path.exists(f)]
    if not valid_files:
        print("Error: None of the specified input files exist.")
        return False
    
    # Create a new Excel writer
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    
    # Track which sheets we've seen
    processed_sheets = set()
    
    # Track scenarios found
    scenarios_found = []
    
    # Store fuel consumption data for later cost calculation
    fuel_consumption_data = {}
    
    # Process each input file
    for file_idx, file_path in enumerate(input_files):
        if not os.path.exists(file_path):
            print(f"Warning: File {file_path} does not exist. Skipping...")
            continue
            
        # Extract scenario name from filename (e.g., 'g1' from 'results_g1.xlsx')
        scenario = os.path.basename(file_path).split('.')[0].split('_')[-1]
        scenarios_found.append(scenario)
        print(f"Processing {file_path} (Scenario: {scenario})...")
        
        try:
            # Read each sheet from the current Excel file
            excel_file = pd.ExcelFile(file_path)
            
            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    
                    # Store fuel consumption data if this is the relevant sheet
                    if sheet_name == 'Fuel_Consumption' and calculate_fuel_costs:
                        fuel_consumption_data[scenario] = df.copy()
                    
                    # Add a scenario column to identify the source file
                    df['Scenario'] = scenario
                    
                    # Move Scenario column to the beginning
                    cols = df.columns.tolist()
                    cols = cols[-1:] + cols[:-1]
                    df = df[cols]
                    
                    # If we've already processed this sheet type, append to existing sheet
                    merged_sheet_name = sheet_name
                    if sheet_name in processed_sheets:
                        # Read existing data
                        existing_df = pd.read_excel(output_file, sheet_name=merged_sheet_name)
                        # Combine with new data
                        combined_df = pd.concat([existing_df, df], ignore_index=True)
                        # Write back to the same sheet
                        combined_df.to_excel(writer, sheet_name=merged_sheet_name, index=False)
                    else:
                        # First time seeing this sheet, just write it
                        df.to_excel(writer, sheet_name=merged_sheet_name, index=False)
                        processed_sheets.add(sheet_name)
                except Exception as e:
                    print(f"Error processing sheet {sheet_name} in {file_path}: {str(e)}")
        except Exception as e:
            print(f"Error processing file {file_path}: {str(e)}")
    
    # Calculate and add fuel costs if requested
    if calculate_fuel_costs and fuel_consumption_data:
        try:
            fuel_costs_df = calculate_fuel_costs(fuel_consumption_data, scenarios_found)
            if not fuel_costs_df.empty:
                fuel_costs_df.to_excel(writer, sheet_name='Fuel_Costs', index=False)
                processed_sheets.add('Fuel_Costs')
                print("Added Fuel_Costs sheet to the merged file")
        except Exception as e:
            print(f"Error calculating fuel costs: {str(e)}")
    
    # Save the output file
    writer.close()
    print(f"Merged data saved to {output_file}")
    
    # Add color-coding and formatting
    if add_colors:
        add_color_coding(output_file, scenarios_found)
    
    # Create summary sheet with charts
    create_summary_charts(output_file, scenarios_found)
    
    return True

def calculate_fuel_costs(fuel_consumption_data, scenarios):
    """
    Calculate fuel costs based on fuel consumption data and standard fuel prices.
    
    Parameters:
    fuel_consumption_data (dict): Dictionary of DataFrames with fuel consumption by scenario
    scenarios (list): List of scenarios
    
    Returns:
    DataFrame: Calculated fuel costs
    """
    # Define standard fuel prices ($/unit)
    # These are example prices - adjust as needed for your specific fuels
    fuel_prices = {
        'coal': 100,       # $/ton
        'natural_gas': 5,  # $/MMBtu
        'electricity': 70, # $/MWh
        'hydrogen': 5,     # $/kg
        'biomass': 120,    # $/ton
        'oil': 80,         # $/barrel
        'coke': 300,       # $/ton
        'COG': 4,          # $/MMBtu
        'BFG': 2,          # $/MMBtu
        'BOFG': 3,         # $/MMBtu
    }
    
    # Initialize results list
    fuel_costs_data = []
    
    # Process each scenario
    for scenario in scenarios:
        if scenario not in fuel_consumption_data:
            continue
            
        df = fuel_consumption_data[scenario]
        
        # Process each row in the fuel consumption data
        for _, row in df.iterrows():
            system = row.get('System', 'Unknown')
            year = row.get('Year', 0)
            fuel = row.get('Fuel', '')
            consumption = row.get('Consumption', 0)
            
            # Calculate cost based on fuel type and consumption
            fuel_price = fuel_prices.get(fuel.lower(), 0)  # Default to 0 if fuel not found
            if fuel_price > 0 and consumption > 0:
                cost = fuel_price * consumption
                
                fuel_costs_data.append({
                    'Scenario': scenario,
                    'System': system,
                    'Year': year,
                    'Fuel': fuel,
                    'Consumption': consumption,
                    'Unit_Price': fuel_price,
                    'Total_Cost': cost
                })
    
    # Create DataFrame from results
    if fuel_costs_data:
        fuel_costs_df = pd.DataFrame(fuel_costs_data)
        
        # Also create a summary by scenario and year
        summary_data = []
        for scenario in scenarios:
            scenario_data = fuel_costs_df[fuel_costs_df['Scenario'] == scenario]
            for year in scenario_data['Year'].unique():
                year_data = scenario_data[scenario_data['Year'] == year]
                total_cost = year_data['Total_Cost'].sum()
                
                summary_data.append({
                    'Scenario': scenario,
                    'Year': year,
                    'Total_Fuel_Cost': total_cost
                })
        
        # Append summary to the bottom with a separator row
        if summary_data:
            summary_df = pd.DataFrame(summary_data)
            separator = pd.DataFrame([{col: '' for col in fuel_costs_df.columns}])
            summary_header = pd.DataFrame([{'Scenario': 'SUMMARY', 'System': '', 'Year': '', 'Fuel': '', 
                                           'Consumption': '', 'Unit_Price': '', 'Total_Cost': ''}])
            
            # Add columns to summary_df to match fuel_costs_df
            for col in fuel_costs_df.columns:
                if col not in summary_df.columns:
                    summary_df[col] = ''
            
            # Ensure Total_Cost column exists in summary_df
            if 'Total_Cost' in fuel_costs_df.columns and 'Total_Fuel_Cost' in summary_df.columns:
                summary_df['Total_Cost'] = summary_df['Total_Fuel_Cost']
                summary_df = summary_df.drop(columns=['Total_Fuel_Cost'])
            
            fuel_costs_df = pd.concat([fuel_costs_df, separator, summary_header, summary_df], ignore_index=True)
        
        return fuel_costs_df
    else:
        return pd.DataFrame()  # Return empty DataFrame if no data

def add_color_coding(excel_file, scenarios=None):
    """
    Add color-coding to an Excel file based on scenario values.
    
    Parameters:
    excel_file (str): Path to the Excel file to modify
    scenarios (list): List of scenarios found in the data
    """
    # Define colors for different scenarios (can handle more scenarios)
    scenario_colors = {
        'g1': 'FFCCFFCC',  # Light green
        'g2': 'FFFFCCCC',  # Light red/pink
        'g3': 'FFCCCCFF',  # Light blue
        'g4': 'FFFFF2CC',  # Light yellow
        'g5': 'FFE6E6E6',  # Light gray
        'g6': 'FFD9D2E9',  # Light purple
        'g7': 'FFDDEEFF',  # Light cyan
        'g8': 'FFFCE5CD',  # Light orange
    }
    
    # Define header style
    header_fill = PatternFill(start_color="FF366092", end_color="FF366092", fill_type='solid')
    header_font = Font(color="FFFFFFFF", bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Define summary style
    summary_fill = PatternFill(start_color="FFD8E4BC", end_color="FFD8E4BC", fill_type='solid')
    summary_font = Font(bold=True)
    
    # Load the workbook
    try:
        wb = load_workbook(excel_file)
        
        # Process each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format headers
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) if max_length < 50 else 50
                ws.column_dimensions[column].width = adjusted_width
            
            # Find the scenario column index
            scenario_col = None
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value == 'Scenario':
                    scenario_col = col_idx
                    break
            
            if scenario_col is None:
                continue  # Skip if no Scenario column
                
            # Apply colors based on scenario value and handle summary rows
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                cell = row[scenario_col-1]  # Convert to 0-based index
                
                # Check if this is a summary row
                if cell.value == 'SUMMARY':
                    # Apply summary styling
                    for c in row:
                        c.fill = summary_fill
                        c.font = summary_font
                        c.border = thin_border
                elif cell.value in scenario_colors:
                    # Apply scenario color
                    fill = PatternFill(start_color=scenario_colors[cell.value], 
                                      end_color=scenario_colors[cell.value], 
                                      fill_type='solid')
                    for c in row:
                        c.fill = fill
                        c.border = thin_border
        
        # Save the workbook
        wb.save(excel_file)
        print(f"Added color coding and formatting to {excel_file}")
    except Exception as e:
        print(f"Error adding color coding: {str(e)}")

def create_summary_charts(excel_file, scenarios):
    """
    Create a summary sheet with charts comparing key metrics across scenarios.
    
    Parameters:
    excel_file (str): Path to the Excel file
    scenarios (list): List of scenarios to include in the charts
    """
    try:
        # Load the workbook
        wb = load_workbook(excel_file)
        
        # Check if Global_Summary sheet exists
        if 'Global_Summary' not in wb.sheetnames:
            print("No Global_Summary sheet found. Skipping chart creation.")
            return
        
        # Create a new summary sheet
        if 'Scenario_Comparison' in wb.sheetnames:
            ws_summary = wb['Scenario_Comparison']
            # Clear existing content
            for row in ws_summary.iter_rows():
                for cell in row:
                    cell.value = None
        else:
            ws_summary = wb.create_sheet('Scenario_Comparison', 0)
        
        # Set title
        ws_summary['A1'] = 'Scenario Comparison Charts'
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary.merge_cells('A1:H1')
        
        # Read Global_Summary data
        ws_global = wb['Global_Summary']
        
        # Find key metrics columns
        header_row = next(ws_global.rows)
        col_indices = {}
        for idx, cell in enumerate(header_row):
            col_indices[cell.value] = idx
        
        # Key metrics to chart
        metrics = [
            'Total_Emissions', 
            'Total_Production', 
            'Global_Emission_Intensity',
            'Total_CAPEX',
            'Total_OPEX',
            'Total_Cost'
        ]
        
        # Add fuel cost if available
        if 'Fuel_Costs' in wb.sheetnames:
            metrics.append('Total_Fuel_Cost')
            
            # Try to extract fuel cost data
            try:
                ws_fuel_costs = wb['Fuel_Costs']
                fuel_cost_data = {}
                
                # Find the summary section
                in_summary = False
                for row in ws_fuel_costs.iter_rows(min_row=1):
                    if row[0].value == 'SUMMARY':
                        in_summary = True
                        continue
                    
                    if in_summary and row[0].value in scenarios:
                        scenario = row[0].value
                        year = row[2].value  # Assuming Year is in column C
                        cost = row[6].value  # Assuming Total_Cost is in column G
                        
                        if scenario not in fuel_cost_data:
                            fuel_cost_data[scenario] = {}
                        
                        if year is not None and cost is not None:
                            fuel_cost_data[scenario][year] = cost
                
                # Add fuel cost data to Global_Summary for charting
                if fuel_cost_data:
                    # Find the last column to add fuel cost
                    last_col = len(header_row) + 1
                    ws_global.cell(row=1, column=last_col, value='Total_Fuel_Cost')
                    
                    # Add data for each row
                    for row_idx, row in enumerate(ws_global.iter_rows(min_row=2), 2):
                        scenario = row[col_indices['Scenario']].value
                        year = row[col_indices['Year']].value
                        
                        if scenario in fuel_cost_data and year in fuel_cost_data[scenario]:
                            ws_global.cell(row=row_idx, column=last_col, value=fuel_cost_data[scenario][year])
                    
                    # Update col_indices
                    col_indices['Total_Fuel_Cost'] = last_col - 1
            except Exception as e:
                print(f"Error adding fuel cost data to charts: {str(e)}")
        
        # Create data for each metric
        row_offset = 3
        for metric_idx, metric in enumerate(metrics):
            if metric not in col_indices:
                continue
                
            # Add metric title
            title_row = row_offset + metric_idx * 20
            ws_summary.cell(row=title_row, column=1, value=f"{metric} by Scenario and Year")
            ws_summary.cell(row=title_row, column=1).font = Font(size=14, bold=True)
            ws_summary.merge_cells(f'A{title_row}:H{title_row}')
            
            # Add headers
            ws_summary.cell(row=title_row+1, column=1, value="Year")
            for scen_idx, scenario in enumerate(scenarios):
                ws_summary.cell(row=title_row+1, column=2+scen_idx, value=scenario)
            
            # Extract data
            years = set()
            data = {scenario: {} for scenario in scenarios}
            
            for row in ws_global.iter_rows(min_row=2):
                year = row[col_indices['Year']].value
                scenario = row[col_indices['Scenario']].value
                if scenario in scenarios and metric in col_indices:
                    value = row[col_indices[metric]].value
                    if value is not None:
                        years.add(year)
                        data[scenario][year] = value
            
            # Sort years
            years = sorted(years)
            
            # Write data
            for year_idx, year in enumerate(years):
                ws_summary.cell(row=title_row+2+year_idx, column=1, value=year)
                for scen_idx, scenario in enumerate(scenarios):
                    if year in data[scenario]:
                        ws_summary.cell(row=title_row+2+year_idx, column=2+scen_idx, value=data[scenario][year])
            
            # Create chart
            chart = LineChart()
            chart.title = f"{metric} by Year and Scenario"
            chart.style = 2
            chart.x_axis.title = "Year"
            chart.y_axis.title = metric.replace('_', ' ')
            
            data_ref = Reference(ws_summary, min_row=title_row+1, max_row=title_row+1+len(years),
                                min_col=1, max_col=1+len(scenarios))
            cats_ref = Reference(ws_summary, min_row=title_row+2, max_row=title_row+1+len(years),
                                min_col=1, max_col=1)
            
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            
            # Place chart
            ws_summary.add_chart(chart, f"J{title_row}")
        
        # Save workbook
        wb.save(excel_file)
        print(f"Created summary charts in {excel_file}")
    except Exception as e:
        print(f"Error creating summary charts: {str(e)}")

def compare_scenarios(input_files, output_file, metrics_to_compare=None):
    """
    Create a focused comparison of specific metrics across scenarios.
    
    Parameters:
    input_files (list): List of input Excel file paths
    output_file (str): Path for the output Excel file
    metrics_to_compare (list): List of metrics to compare (if None, uses default key metrics)
    """
    if metrics_to_compare is None:
        metrics_to_compare = [
            'Total_Emissions', 
            'Total_Production', 
            'Global_Emission_Intensity',
            'Total_CAPEX',
            'Total_OPEX',
            'Total_Cost',
            'Total_Fuel_Cost'  # Added fuel cost
        ]
    
    # First merge the files
    if not merge_excel_files(input_files, output_file):
        return
    
    # Now create a focused comparison sheet
    try:
        # Load the workbook
        wb = load_workbook(output_file)
        
        # Check if Global_Summary sheet exists
        if 'Global_Summary' not in wb.sheetnames:
            print("No Global_Summary sheet found. Cannot create comparison.")
            return
        
        # Create a new comparison sheet
        if 'Metric_Comparison' in wb.sheetnames:
            ws_compare = wb['Metric_Comparison']
            # Clear existing content
            for row in ws_compare.iter_rows():
                for cell in row:
                    cell.value = None
        else:
            ws_compare = wb.create_sheet('Metric_Comparison')
        
        # Set title
        ws_compare['A1'] = 'Scenario Metric Comparison'
        ws_compare['A1'].font = Font(size=16, bold=True)
        ws_compare.merge_cells('A1:F1')
        
        # Read Global_Summary data
        df = pd.read_excel(output_file, sheet_name='Global_Summary')
        
        # If we have fuel costs, try to add them to the comparison
        if 'Fuel_Costs' in wb.sheetnames and 'Total_Fuel_Cost' not in df.columns:
            try:
                fuel_costs_df = pd.read_excel(output_file, sheet_name='Fuel_Costs')
                
                # Extract summary data
                summary_data = fuel_costs_df[fuel_costs_df['Scenario'] != '']
                summary_data = summary_data[summary_data['Scenario'] != 'SUMMARY']
                
                # Create a pivot table of fuel costs by scenario and year
                if 'Total_Cost' in summary_data.columns:
                    fuel_cost_pivot = summary_data.pivot_table(
                        values='Total_Cost', 
                        index=['Scenario', 'Year'], 
                        aggfunc='sum'
                    ).reset_index()
                    
                    # Rename column
                    fuel_cost_pivot.rename(columns={'Total_Cost': 'Total_Fuel_Cost'}, inplace=True)
                    
                    # Merge with the Global_Summary data
                    df = pd.merge(
                        df, 
                        fuel_cost_pivot,
                        on=['Scenario', 'Year'],
                        how='left'
                    )
                    
                    # Fill NaN values with 0
                    df['Total_Fuel_Cost'].fillna(0, inplace=True)
            except Exception as e:
                print(f"Error adding fuel costs to comparison: {str(e)}")
        
        # Get unique scenarios and years
        scenarios = df['Scenario'].unique().tolist()
        years = df['Year'].unique().tolist()
        years.sort()
        
        # For each metric, create a comparison table
        row_offset = 3
        for metric_idx, metric in enumerate(metrics_to_compare):
            if metric not in df.columns:
                continue
                
            # Add metric title
            title_row = row_offset + metric_idx * (len(years) + 5)
            ws_compare.cell(row=title_row, column=1, value=f"{metric.replace('_', ' ')} Comparison")
            ws_compare.cell(row=title_row, column=1).font = Font(size=14, bold=True)
            ws_compare.merge_cells(f'A{title_row}:F{title_row}')
            
            # Add headers
            ws_compare.cell(row=title_row+1, column=1, value="Year")
            for scen_idx, scenario in enumerate(scenarios):
                ws_compare.cell(row=title_row+1, column=2+scen_idx, value=scenario)
            
            # Add data rows
            for year_idx, year in enumerate(years):
                ws_compare.cell(row=title_row+2+year_idx, column=1, value=year)
                
                for scen_idx, scenario in enumerate(scenarios):
                    # Get value for this scenario and year
                    value = df[(df['Scenario'] == scenario) & (df['Year'] == year)][metric].values
                    if len(value) > 0:
                        ws_compare.cell(row=title_row+2+year_idx, column=2+scen_idx, value=value[0])
            
            # Add percent difference row (compared to first scenario)
            if len(scenarios) > 1:
                diff_row = title_row+2+len(years)
                ws_compare.cell(row=diff_row, column=1, value="% Diff from " + scenarios[0])
                
                for scen_idx, scenario in enumerate(scenarios[1:], 1):
                    # Calculate average percent difference across years
                    diffs = []
                    for year_idx, year in enumerate(years):
                        base_val = df[(df['Scenario'] == scenarios[0]) & (df['Year'] == year)][metric].values
                        comp_val = df[(df['Scenario'] == scenario) & (df['Year'] == year)][metric].values
                        
                        if len(base_val) > 0 and len(comp_val) > 0 and base_val[0] != 0:
                            diff_pct = (comp_val[0] - base_val[0]) / base_val[0] * 100
                            diffs.append(diff_pct)
                    
                    if diffs:
                        avg_diff = sum(diffs) / len(diffs)
                        ws_compare.cell(row=diff_row, column=2+scen_idx, value=avg_diff/100)  # Convert to decimal for percentage format
                        # Format as percentage
                        ws_compare.cell(row=diff_row, column=2+scen_idx).number_format = '0.00%'
            
            # Create chart
            chart = BarChart()
            chart.type = "col"
            chart.style = 10
            chart.title = f"{metric.replace('_', ' ')} by Scenario"
            chart.y_axis.title = metric.replace('_', ' ')
            chart.x_axis.title = "Year"
            
            data_len = len(years)
            data = Reference(ws_compare, min_row=title_row+1, max_row=title_row+1+data_len,
                           min_col=2, max_col=1+len(scenarios))
            cats = Reference(ws_compare, min_row=title_row+2, max_row=title_row+1+data_len,
                           min_col=1, max_col=1)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            # Place chart
            ws_compare.add_chart(chart, f"H{title_row}")
        
        # Save workbook
        wb.save(output_file)
        print(f"Created metric comparison in {output_file}")
    except Exception as e:
        print(f"Error creating metric comparison: {str(e)}")

if __name__ == "__main__":
    # Input files to merge
    input_files = [
        'results_g1.xlsx',
        'results_g2.xlsx',
        'results_g3.xlsx'
    ]
    
    # Output file
    output_file = 'merged_results.xlsx'
    
    # Merge the files with fuel cost calculation
    merge_excel_files(input_files, output_file, calculate_fuel_costs=True)
    
    # Create a focused comparison of key metrics including fuel costs
    key_metrics = [
        'Total_Emissions', 
        'Total_Production', 
        'Global_Emission_Intensity',
        'Total_CAPEX',
        'Total_OPEX',
        'Total_Cost',
        'Total_Fuel_Cost'
    ]
    
    compare_scenarios(input_files, 'scenario_comparison.xlsx', key_metrics)
